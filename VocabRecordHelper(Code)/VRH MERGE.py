# encoding: UTF-8
# Created by Abyss_Seeker!
# Modules used: built-in: [tkinter, os, json]; others: [requests, openpyxl]
# Glhf
# This is a file that merges the other 3 code. I used this simply for easier packaging into .exe

import requests
import json
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Side, Border, Font
import tkinter as tk


# From translate_button
def translate(word):
        
    # Global Variable Headers
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.81 Safari/537.36 Edg/104.0.1293.54'}

    # Main body - searching word
    def search(word):
        def org(res_dict):
            if 'entries' in res_dict['data']:
                explanation = res_dict['data']['entries'][0]
                if explanation['entry'] == word:
                    definition = explanation['explain']
                    return '; '.join(definition.split('；')[:3])
            return 'Word not found. Make sure it is in present tense and not spelled wrong! Try again :)'

        word = word.strip().lower()
        # 有道为动态加载，此链接可以从Network - XHR中找到，请求方式为requests
        url = 'https://dict.youdao.com/suggest?num=1&ver=3.0&doctype=json&cache=false&le=en&q={}'.format(word)
        res = requests.get(url, headers = headers)
        # 将获取的信息转为dict格式
        res_dict = res.json()
        definition = org(res_dict)
        return definition

    # Using search function, get the translation
    translation = search(word)
    # Editing output textbox
    output_text.delete('1.0', 'end')
    output_text.insert('1.0', translation)
    # Modify event log label
    # 此处'end-1c'是为了去掉文本框返回时自带的一个\n
    if output_text.get('1.0','end-1c') != 'Word not found. Make sure it is in present tense and not spelled wrong! Try again :)':
        action_label['text'] = 'Translation for word "{}" successly found and presented'.format(word)
    else:
        action_label['text'] = 'Translation for word "{}" not found'.format(word)

# From record_button
def record(word, translation):
    # 加行
    def add_line(word, translation):
        # 定义对齐样式横向靠左、纵向居中
        align = Alignment(horizontal='left', vertical='center')
        new_workbook = False
        # 定义边样式为细条
        side = Side('thin')
        # 定义表头边框样式，有底边和右边
        header_border = Border(bottom=side, right=side)
        # 定义表中、表尾边框样式，有左边
        content_border = Border(left=side)
        
        if os.path.exists('./Recorded vocab list.xlsx') == False:
            # 要是不存在，新建
            wb = Workbook()
            new_workbook = True
        else:
            # 要是存在，不管
            wb = load_workbook('./Recorded vocab list.xlsx')
        ws = wb.active
        # 调整列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 80
        # 要是是新建的，加表头
        if new_workbook:
            ws.append(['Vocabulary', 'Translation'])
            for header_cell in ws[1]:
                header_cell.border = header_border
                header_cell.alignment = align
                header_cell.font = Font(bold=True)

        # 加信息
        ws.append([word,translation])

        # 获取刚加的行
        row_num = ws.max_row
        
        # 加格式
        for cell in ws[row_num]:
            cell.border = content_border
            cell.alignment = align

        # 保存
        wb.save('./Recorded vocab list.xlsx')

        #返回刚加的行数
        return row_num


    row_num = add_line(word, translation)
    action_label['text'] = 'Row {} added: {} - {}'.format(row_num, word, translation)

def delete():
    # 删行    
    def del_line():
        # 要是文件不存在，直接返回
        if os.path.exists('./Recorded vocab list.xlsx') == False:
            return 'File doesn\'t exist'
        else:
            wb = load_workbook('./Recorded vocab list.xlsx')
            ws = wb.active
            row_num = ws.max_row
            # 前提文件行数大于一（不是只有一个headers）
            if row_num != 1:
                word = ws[row_num][0].value
                translation = ws[row_num][1].value
                # 删除最后一行
                ws.delete_rows(row_num)
                # 保存
                wb.save('./Recorded vocab list.xlsx')
                # 返回信息
                return {'word':word, 'translation':translation, 'row_num':row_num}
            else:
                return 'No lines left'
    line_info = del_line()
    if line_info == 'File doesn\'t exist' or line_info == 'No lines left':
        action_label['text'] = 'Delete line failed: {}'.format(line_info)
    else:
        action_label['text'] = 'Row {} successfully deleted: {} - {}'.format(line_info['row_num'], line_info['word'], line_info['translation'])


# 说明窗口
def help_window():
    help_root = tk.Tk()
    help_root.title('How to use Vocab Record Helper')
    help_label = tk.Label(help_root, text = '''Hi! I am nothing but a guide for you to know how to use this recorder... You can close me if it's not your first time using it

    The whole thing is really basic and easy: you can enter the vocab you don't know as instructed, then click the 'Translate' button, and the translation would appear :)

    If you are satisfied with the translation, you can just go foward and click 'record' to record the vocab and the translation down in an excel sheet!

    You may also do small modifications to the translation or whatever you want!

    At last, if you made a mistake or something, you can click the 'delete' button. This deletes the last entry you made in the excel! Don't click too many times!

    I'm still kinda noob in python so there may be lotsss of bugs. You are welcomed to inform me of that or alter the code yourself!

    You can always find me at the top right corner :D

    Have fun! Now you may close the tab :P

    From Abyss_Seeker!
    ''')
    help_label.pack()

# Show help window
help_window()

# User interface
root = tk.Tk()
root.configure(bg='#EE51B1')
root.geometry('500x450+200+200')
root.resizable(False, False)
root.title('Vocab Record Helper - User Interface')

# Setting title
title_label = tk.Label(root, text='Vocab Record Helper', fg='white', bg='#E3C515', font=('Arial',18), height=2, width=40)
title_label.place(x=-30,y=0)

# Initializing input label and entry box
input_label = tk.Label(root, text='Please enter the vocab:', fg='white', bg='#EE51B1', font=('Arial', 14))
input_label.place(x=20, y=80)
input_entry = tk.Entry(root, fg='white', bg='#4B2D9F', font=('Arial', 14), width=18)
input_entry.insert(0, 'Example')
input_entry.place(x=250, y=80)

# Initializing output label and text box
output_label = tk.Label(root, text='Translation:', fg='white', bg='#EE51B1', font=('Arial', 14))
output_label.place(x=20, y=120)
output_text = tk.Text(root, fg='white', bg='#4B2D9F', font=('Arial', 14), width=41, height=6)
output_text.insert('1.0','The translation will appear here.')
output_text.place(x=20, y=160)

# Initializing left button, when button clicked, function "translate" will be triggered
translate_button = tk.Button(root, text='Translate', fg='white', bg='#2C5D37', font=('Arial', 14, 'bold'), width = 10, height = 2,
                   command=lambda: translate(input_entry.get()))
translate_button.place(x=20, y=320)

# Initializing middle button, when button clicked, function "record" will be triggered
record_button = tk.Button(root, text='Record', fg='white', bg='#E3C515', font=('Arial', 14, 'bold'), width = 10, height = 2,
                   command=lambda: record(input_entry.get().lower(), output_text.get('1.0','end-1c')))
record_button.place(x=180, y=320)

# Initializing right button, when button clicked, function "delete" will be triggered
record_button = tk.Button(root, text='Delete', fg='white', bg='#A59CD3', font=('Arial', 14, 'bold'), width = 10, height = 2,
                   command=delete)
record_button.place(x=340, y=320)

# Add small comments
cmt_label = tk.Label(root, text='*While using Record and Delete, make sure the excel file is closed!', fg='white', bg='#EE51B1', font=('Arial', 10))
cmt_label.place(x=20, y=385)

name_label = tk.Label(root, text='@Abyss_Seeker!', fg='#BAFFBA', bg='#EE51B1', font=('Arial', 6))
name_label.place(x=440,y=435)

# Initializing event log (title and log) at bottom
event_log_label = tk.Label(root, text='Event log: Recent actions will appear below:', fg='white', bg='#EE51B1', font=('Arial', 8))
event_log_label.place(x=20, y=410)
action_label = tk.Label(root, text='No events yet', fg='white', bg='#EE51B1', font=('Arial', 8))
action_label.place(x=20, y=428)

# Help button, shows help window
record_button = tk.Button(root, text='?', fg='#BAFFBA', bg='#E3C515', font=('Arial', 8, 'bold'), width = 2, height = 1,
                   command=help_window)
record_button.place(x=475, y=0)

# Let's rock
root.mainloop()



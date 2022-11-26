import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Side, Border, Font

# 加行
def add_line(word, translation):
    # 定义对齐样式横向靠左、纵向居中
    align = Alignment(horizontal='left', vertical='center')
    new_workbook = False
    # 定义边样式为细条
    side = Side('thin')
    # 定义表头边框样式，有底边和右边
    header_border = Border(bottom=side, right=side)
    # 定义表中、表尾边框样式，有左边
    content_border = Border(left=side)
    
    if os.path.exists('./Recorded vocab list.xlsx') == False:
        # 要是不存在，新建
        wb = Workbook()
        new_workbook = True
    else:
        # 要是存在，不管
        wb = load_workbook('./Recorded vocab list.xlsx')
    ws = wb.active
    # 调整列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 80
    # 要是是新建的，加表头
    if new_workbook:
        ws.append(['Vocabulary', 'Translation'])
        for header_cell in ws[1]:
            header_cell.border = header_border
            header_cell.alignment = align
            header_cell.font = Font(bold=True)

    # 加信息
    ws.append([word,translation])

    # 获取刚加的行
    row_num = ws.max_row
    
    # 加格式
    for cell in ws[row_num]:
        cell.border = content_border
        cell.alignment = align

    # 保存
    wb.save('./Recorded vocab list.xlsx')

    #返回刚加的行数
    return row_num

# 删行    
def del_line():
    # 要是文件不存在，直接返回
    if os.path.exists('./Recorded vocab list.xlsx') == False:
        return 'File doesn\'t exist'
    else:
        wb = load_workbook('./Recorded vocab list.xlsx')
        ws = wb.active
        row_num = ws.max_row
        # 前提文件行数大于一（不是只有一个headers）
        if row_num != 1:
            word = ws[row_num][0].value
            translation = ws[row_num][1].value
            # 删除最后一行
            ws.delete_rows(row_num)
            # 保存
            wb.save('./Recorded vocab list.xlsx')
            # 返回信息
            return {'word':word, 'translation':translation, 'row_num':row_num}
        else:
            return 'No lines left'


            
